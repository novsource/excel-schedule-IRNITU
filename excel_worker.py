import re
import json
import os
from operator import attrgetter
from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook

pairs = {
    '8.15-9.45': 1,
    '9.55-11.25': 2,
    '12.05-13.35': 3,
    '13.45-15.15': 4,
    '15.25-16.55': 5,
    '17.05-18.35': 6
}

days = {
    'понедельник': 1,
    'вторник': 2,
    'среда': 3,
    'четверг': 4,
    'пятница': 5,
    'суббота': 6
}


def get_started():
    path = input('Введите путь к файлу: ')

    if is_xls(path) is True:
        path = convert_to_xlsx(path)

    book = load_workbook(path)

    for sheet_name in book.sheetnames:
        if sheet_name != 'аудитории':
            worksheet = book[sheet_name]
            excel_into_json(worksheet)


def convert_to_xlsx(path_to_file):
    print('Файл с расширением .xls не поддерживается поэтому он будет конвертирован в файл с расширением .xlsx')
    x2x = XLS2XLSX(path_to_file)
    x2x.to_xlsx(path_to_file + 'x')
    path = path_to_file+'x'
    print('Файл успешно конвертирован к формату .xlsx!\n')
    return path


def is_xls(path_to_file):
    name = os.path.basename(path_to_file)
    name = str(name).split('.')
    if name[1] == 'xls':
        return True
    else:
        return False


def get_cell_of_beginning_table(worksheet):
    beg_table = {}
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == 'Учебная группа':
                beg_table[worksheet.title] = cell
    return beg_table


def get_students_group_from_sheet(worksheet):
    beg_table = int(get_cell_of_beginning_table(worksheet).get(worksheet.title).row)  # Получаем начало таблицы из листа

    groups_cells = {}  # Словарь аудиторий с клетками

    for row in worksheet.iter_rows(min_row=beg_table, max_row=beg_table, max_col=worksheet.max_column):
        for cell in row:
            if (cell.value is not None) and (cell.value != 'Учебная группа'):
                groups_cells[cell.value] = cell

    return groups_cells


def get_cells_schedule(worksheet):
    beg_table = {}
    flag = False
    for row in worksheet.iter_rows(min_row=1, max_col=1):
        for cell in row:
            value = str(cell.value).replace(" ", '')
            if days.get(value) == 1:
                beg_table['Begin'] = cell
            if days.get(value) == 6:
                flag = True
            if flag is True and pairs.get(str(worksheet.cell(row[0].row, 2).value).replace(' ','')) == 6:
                beg_table['End'] = worksheet.cell(cell.row+1, cell.column)
    return beg_table


def printDataFromGroup(worksheet):
    json_file_out = {'pairs': pairs, 'schedule': get_students_group_from_sheet(worksheet)}
    items = get_students_group_from_sheet(worksheet).items()
    for row in range(1, worksheet.max_row):
        for col in range(1, worksheet.max_column):
            cell = worksheet.cell(row=row, column=col)
            if cell.value is not None:
                print(cell.value)


def excel_into_json(worksheet):
    print('Начало записи json')

    day = 0
    time_pair = 0

    groups_students = get_students_group_from_sheet(worksheet)
    groups = dict((k, {}) for k in list(groups_students.keys()))

    json_out = {'pairs': {v: k for k, v in pairs.items()}, 'schedule': groups}

    for group in groups:
        json_2 = {}
        dict_in_json = {}

        group_cell = groups_students.get(group)

        for row in range(get_cells_schedule(worksheet).get('Begin').row, get_cells_schedule(worksheet).get('End').row):

            if worksheet.cell(row, 1).value is not None:
                into = dict(k for k in dict_in_json.items())

                if len(into) != 0:
                    json_2.update({day: into})
                    dict_in_json.clear()

                day = str(days.get(str(worksheet.cell(row, 1).value).replace(' ', '')))

            if worksheet.cell(row, 2).value is not None:
                time_pair = str(pairs.get(str(worksheet.cell(row, 2).value).replace(' ', '')))

            for col in range(group_cell.column, group_cell.column + 1):
                time = worksheet.cell(row, 2)
                cell = worksheet.cell(row, col)

                if is_empty(cell) is False:
                    pair_week = get_pair_week(worksheet, time, cell)

                    if dict_in_json.get(time_pair) is None and (pair_week is not None):
                        dict_in_json.update({time_pair: pair_week})
                    else:
                        if (pair_week is not None) and dict_in_json.get(time_pair) is not None:
                            dict_in_json[time_pair].append(pair_week[0])

        if len(json_2) != 0:
            into = dict(k for k in dict_in_json.items())
            json_2.update({day: into})
            json_out['schedule'].update({group: json_2})

    with open('result.json', 'w', encoding='utf-8') as fp:
        json.dump(json_out, fp, indent=4, ensure_ascii=False)
        print('JSON-файл готов')


def get_pair_week(worksheet, time, pair):
    list_out = []

    title = get_pair_title(pair)
    teachers_list = get_teachers(pair)

    if len(teachers_list) == 0:
        teachers_list = ['']

    audit_list = get_audit(worksheet, pair)

    if is_merged(worksheet, pair) and pair.value is not None:
        if len(teachers_list) <= 1 and len(audit_list) == 2:  # Физкультура
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 1,
                         'aud': [audit_list[0], audit_list[1]]}
            list_out.append(pair_dict)
            return list_out
        if len(teachers_list) > 1:
            for i in range(len(teachers_list)):
                pair_dict = {'title': title,
                             'teacher': teachers_list[i],
                             'week': 0,
                             'aud': [audit_list[i]]}
                list_out.append(pair_dict)
        else:
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 0,
                         'aud': [audit_list[0]]}
            list_out.append(pair_dict)
        return list_out
    if (time.value is not None) and (pair.value is not None):
        if len(teachers_list) <= 1 and len(audit_list) == 2:  # Физкультура
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 1,
                         'aud': [audit_list[0], audit_list[1]]}
            list_out.append(pair_dict)
            return list_out
        if len(teachers_list) > 1:
            for i in range(len(teachers_list)):
                pair_dict = {'title': title,
                             'teacher': teachers_list[i],
                             'week': 1,
                             'aud': [audit_list[i]]}
                list_out.append(pair_dict)
        else:
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 1,
                         'aud': [audit_list[0]]}
            list_out.append(pair_dict)
        return list_out
    if (time.value is None) and (pair.value is not None):
        if len(teachers_list) <= 1 and len(audit_list) == 2:  # Физкультура
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 1,
                         'aud': [audit_list[0], audit_list[1]]}
            list_out.append(pair_dict)
            return list_out
        if len(teachers_list) > 1:
            for i in range(len(teachers_list)):
                pair_dict = {'title': title,
                             'teacher': teachers_list[i],
                             'week': 2,
                             'aud': [audit_list[i]]}
                list_out.append(pair_dict)
        else:
            pair_dict = {'title': title,
                         'teacher': teachers_list[0],
                         'week': 2,
                         'aud': [audit_list[0]]}
            list_out.append(pair_dict)
        return list_out
    if (pair.value == time.value is None):
        return None


def get_teachers(pair):
    regex = re.compile(r"[А-Я][а-я]*\s\w[.]\w[.]")
    teacher_list = re.findall(regex, str(pair.value))
    if teacher_list != None:
        return teacher_list


def get_audit(worksheet, pair):
    is_fisk = []
    audit_list = []
    audit_cell = worksheet.cell(pair.row, pair.column + 1)
    if audit_cell.value != None:
        is_fisk = re.findall(r'стадион ИРНИТУ', audit_cell.value)
    if audit_cell.value != None and len(is_fisk) == 0: #  Если это не физкультура
        audit = str(audit_cell.value).replace('\n', ',').replace(' ', ',')
        audit_list = str(audit).split(',')
        audit_list = [aud for aud in audit_list if aud != '']
    if audit_cell.value != None and len(is_fisk) != 0:
        audit_list = str(audit_cell.value).split(',')
    if len(audit_list) != 0:
        return audit_list
    else:
        return ['']


def get_pair_title(pair):
    regex = re.compile(r"[А-Я][а-я]*\s\w[.]\w[.]")
    title = re.split(regex, str(pair.value))[0].replace('\n', '').strip()
    if title != None:
        return title


def is_merged(worksheet, cell):
    for merged_cell in sorted(worksheet.merged_cell_ranges, key=attrgetter('coord')):
        if cell.coordinate in merged_cell:
            return True


def is_empty(cell):
    check = str(cell.value).replace(' ', '').replace('\n', '')
    check = [x for x in check if x != '']
    if len(check) != 0 and cell.value is not None:
        return False
    else:
        return True